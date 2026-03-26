"""
Create an Excel report from validated CAP JSON files using a YAML config.
"""

import json
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .flattener import flatten_json

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_COL_WIDTH = 100  # characters
MAX_ROW_HEIGHT = 50  # points

# Styles
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="medium", color="2F5496"),
    right=Side(style="thin", color="D9E2F3"),
)

_DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_DATA_FONT = Font(size=10)

_ALT_ROW_FILL = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
)


def load_config(config_path: str | Path) -> list[str]:
    """Load the YAML config and return the list of enabled column keys."""
    with open(config_path, encoding="utf-8-sig") as f:
        config = yaml.safe_load(f)
    columns = config.get("columns", {})
    return [key for key, enabled in columns.items() if enabled]


def generate_report(input_dir: str | Path, config_path: str | Path, output_path: str | Path):
    """Read all JSON files, flatten them, and write an Excel workbook."""
    columns = load_config(config_path)
    if not columns:
        print("No columns enabled in config. Nothing to report.")
        return

    input_dir = Path(input_dir)
    json_files = sorted(input_dir.glob("*.json"))

    # Exclude any file that looks like a schema or config
    json_files = [f for f in json_files if "schema" not in f.name.lower()]

    if not json_files:
        print(f"No JSON files found in {input_dir}")
        return

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "CAP Policies"

    headers = ["filename"] + columns

    # Track max content width per column for auto-sizing
    # Seed with header text length (use last dot-segment for readability)
    col_widths = [len(h.split(".")[-1]) for h in headers]
    # First column uses full "filename" header
    col_widths[0] = len(headers[0])

    # --- Header row ---
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER

    # --- Data rows ---
    for row_idx, json_file in enumerate(json_files, 2):
        with open(json_file, encoding="utf-8-sig") as f:
            data = json.load(f)

        flat = flatten_json(data)
        is_alt_row = (row_idx % 2 == 0)

        # Filename column
        cell = ws.cell(row=row_idx, column=1, value=json_file.name)
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGNMENT
        cell.border = _THIN_BORDER
        if is_alt_row:
            cell.fill = _ALT_ROW_FILL
        col_widths[0] = max(col_widths[0], len(json_file.name))

        # Data columns
        for col_idx, key in enumerate(columns, 2):
            value = flat.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DATA_FONT
            cell.alignment = _DATA_ALIGNMENT
            cell.border = _THIN_BORDER
            if is_alt_row:
                cell.fill = _ALT_ROW_FILL

            # Track width: use longest line in multi-line values
            if isinstance(value, str) and value:
                max_line = max(len(line) for line in value.split("\n"))
                col_widths[col_idx - 1] = max(col_widths[col_idx - 1], max_line)
            elif value is not None:
                col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))

        # Auto-adjust row height based on max line count in this row
        max_lines = 1
        for col_idx, key in enumerate(columns, 2):
            value = flat.get(key, "")
            if isinstance(value, str) and "\n" in value:
                max_lines = max(max_lines, value.count("\n") + 1)
        # ~15 points per line, capped at MAX_ROW_HEIGHT
        row_height = min(max(15 * max_lines, 15), MAX_ROW_HEIGHT)
        ws.row_dimensions[row_idx].height = row_height

    # --- Column widths ---
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(col_idx)
        # Fit to content with small padding, cap at MAX_COL_WIDTH
        adjusted = min(width + 2, MAX_COL_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted

    # --- Header row height ---
    ws.row_dimensions[1].height = 30

    # --- Freeze header row and filename column ---
    ws.freeze_panes = "B2"

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Report written to {output_path} ({len(json_files)} policies, {len(columns)} columns)")
